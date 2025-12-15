from django.shortcuts import render
from django.http import JsonResponse
from openpyxl import load_workbook
import io
from rest_framework import viewsets
from .models import Cutting, Size, Material, Color, WorkTable, Reference, WorkShop, CuttingAssignment
from .serializers import (
    CuttingSerializer,
    SizeSerializer,
    MaterialSerializer,
    ColorSerializer,
    WorkTableSerializer,
    ReferenceSerializer,
    WorkShopSerializer,
    CuttingAssignmentSerializer
)


class CuttingViewSet(viewsets.ModelViewSet):
    queryset = Cutting.objects.all()
    serializer_class = CuttingSerializer


class SizeViewSet(viewsets.ModelViewSet):
    queryset = Size.objects.all()
    serializer_class = SizeSerializer


class MaterialViewSet(viewsets.ModelViewSet):
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer


class ColorViewSet(viewsets.ModelViewSet):
    queryset = Color.objects.all()
    serializer_class = ColorSerializer


class WorkTableViewSet(viewsets.ModelViewSet):
    queryset = WorkTable.objects.all()
    serializer_class = WorkTableSerializer


class ReferenceViewSet(viewsets.ModelViewSet):
    queryset = Reference.objects.all()
    serializer_class = ReferenceSerializer


class WorkShopViewSet(viewsets.ModelViewSet):
    queryset = WorkShop.objects.all()
    serializer_class = WorkShopSerializer


class CuttingAssignmentViewSet(viewsets.ModelViewSet):
    queryset = CuttingAssignment.objects.all()
    serializer_class = CuttingAssignmentSerializer

# Excel Upload View
def upload_excel(request):
    if request.method == "POST" and request.FILES['file']:
        excel_file = request.FILES['file']
        try:
            # Leer el archivo Excel
            wb = load_workbook(io.BytesIO(excel_file.read()))
            sheet = wb.active
            # Asumiendo que la primera fila es para encabezados
            rows = sheet.iter_rows(min_row=2, values_only=True)

            # Iterar sobre las filas del archivo Excel
            for row in rows:
                # Asegurarse de que los datos de la fila estén completos
                if len(row) == 6:  # Si hay 6 columnas como se espera (ajusta según sea necesario)
                    reference, qty_color, qty_size, color, size, material = row

                    # Validación básica: asegurarse de que no falten valores
                    if not reference or not qty_color or not qty_size or not color or not size or not material:
                        return JsonResponse({"error": f"Missing data in row {row}"}, status=400)

                    # Crear una nueva entrada en el modelo Cutting
                    Cutting.objects.create(
                        reference=reference,
                        quantity=qty_color,
                        qty_color=qty_size,
                        color=color,
                        size=size,
                        material=material
                    )

            return JsonResponse({"message": "Excel data uploaded successfully"}, status=200)

        except Exception as e:
            return JsonResponse({"error": f"Error processing file: {str(e)}"}, status=500)

    return JsonResponse({"error": "No file uploaded"}, status=400)
